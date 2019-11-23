import openpyxl
import requests
from bs4 import BeautifulSoup
import time
import os
import socket

# 下载医生图片
def download_hospital_icon(path, img_url, img_code):
	try:
		headers = {
			'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36',
			'Host': 'yyk.39.net',
			'Upgrade-Insecure-Requests': '1',
			'Cookie': 'JSESSIONID=abcZ3TdW4CbiabeOULg6w; Hm_lvt_2e44bf94e67d57ced8420d8af730dd64=1574216073; hisHos=22826%2322826; Hm_lpvt_2e44bf94e67d57ced8420d8af730dd64=1574216075',
			'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
			'Accept-Encoding': 'gzip, deflate',
			'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
			'Cache-Control': 'max-age=0',
			'Connection': 'keep-alive'
			}		 
		r = requests.get(img_url, headers=headers)
		print(f'下载{img_code}图片成功！')
		with open(f'{path}/{img_code}.png', 'wb') as f:
			f.write(r.content)
	except Exception as e:
		print(f'下载{img_code}图片失败！')
		print(f'错误代码：{e}')
		with open(f'{path}/errorlog.txt', 'a+', encoding='utf-8') as f:
			f.write(f'错误代码：{e}---下载 {img_url} 图片失败\n')


# 获取excel中的sheet
def get_sheet_object(sheet_index,excel_data):
	# excel_data = openpyxl.load_workbook("doctors_with_intro.xlsx")
	sheet_page_index = int(sheet_index/50) + 1;
	# print(sheet_page_index)

	try:
		return excel_data[str("Sheet"+str(sheet_page_index))]
	except Exception as e:
		excel_sheet = excel_data.create_sheet("Sheet"+str(sheet_page_index))
		title_string = ['编码', '名称', '主页', '等级', '标签', '别称', '联系方式', '地址', '分院名称', '分院主页', '简介', '科研成果', '获奖荣誉', '先进设备'];
		for index in range(len(title_string)):
			excel_sheet.cell(row=1,column=index+1).value = title_string[index]
		return excel_sheet		

# 获取医院详细信息地址
def get_hospital_info(hospital_url):
	# headers = {
	# 	'User-Agent': ' Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36',
	# 	'Host': 'yyk.39.net',
	# 	'Upgrade-Insecure-Requests': '1',
	# 	'Cookie': 'userLikesIdTemp=1572915576702; JSESSIONID=abcmEpnBC68P1eeLfxc6w; Hm_lvt_2e44bf94e67d57ced8420d8af730dd64=1572914529,1572916922,1574145071; hisHos=21022%2322857%231010169%231010169; Hm_lpvt_2e44bf94e67d57ced8420d8af730dd64=1574212618',
	# 	'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
	# 	'Accept-Encoding': 'gzip, deflate',
	# 	'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
	# 	'Cache-Control': 'max-age=0'
	# 	}		 
	# html_content = requests.get(hospital_url, headers=headers)
	# html_content_text = html_content.text.strip()
	# html_content_text = html_content_text.replace('\n','')
	# soup = BeautifulSoup(html_content_text,'lxml')

	# intro_list = soup.select('div[id="keshi_menu_holder"] li')
	# if len(intro_list) > 2:
	# 	html_detail = intro_list[1].a.get('href')
	# 	# print(html_detail)
	# 	return get_hospital_detail(html_detail)
	hospital_url_list1 = hospital_url.split('/')
	hospital_url_list2 = hospital_url_list1[-1].split('.')
	# print(str(hospital_url_list2[0]))
	return	get_hospital_detail('http://yyk.39.net/hospital/' + str(hospital_url_list2[0]) + '_detail.html')

# 获取医院详细信息
def get_hospital_detail(hospital_url):
	headers = {
		'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36',
		'Host': 'yyk.39.net',
		'Upgrade-Insecure-Requests': '1',
		'Cookie': 'JSESSIONID=abcZ3TdW4CbiabeOULg6w; Hm_lvt_2e44bf94e67d57ced8420d8af730dd64=1574216073; hisHos=22826%2322826; Hm_lpvt_2e44bf94e67d57ced8420d8af730dd64=1574216075',
		'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
		'Accept-Encoding': 'gzip, deflate',
		'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
		'Cache-Control': 'max-age=0',
		'Connection': 'keep-alive'
		}		 
	print('get ' + hospital_url)
	html_content = requests.get(hospital_url, headers=headers, timeout=60)
	# html_content = requests.get(hospital_url,timeout=120)
	html_content_text = html_content.text.strip()
	html_content_text = html_content_text.replace('\n','')
	soup = BeautifulSoup(html_content_text,'lxml')


	level = ''	
	level_list = soup.select('div[class="jy_hspt_intro_m"] i[class="pink_link"]')
	if len(level_list) > 0:
		level = level_list[0].get_text()
	# print('level----' + level)

	tag = ''
	tag_list = soup.select('div[class="con hos-intro-sur"] ul[class="sur-tag1 clearfix"]')
	for i in tag_list:
		tag += str(i.get_text() + ' ')
	# print('tag----' + tag)

	nick = ''
	nick_list = soup.select('div[class="jy_hspt_intro_m"] h2')
	if len(nick_list) > 0:
		nick_name = nick_list[0].get_text()
		nick = nick_name[3:] if(len(nick_name)>3) else nick_name
	# print('nick-----' + nick)

	#电话和地址还有分院放一起处理
	phone = ''
	adress = ''
	son_name = ''
	son_url = ''
	info_list = soup.select('dl[class="sur-info clearfix"] dd')
	for i in info_list:
		# print(repr(i.previous_element.previous_element))

		if repr(i.previous_element.previous_element) == "'电话：'":
			phone += str(i.span.get_text() + '\n')
		elif repr(i.previous_element.previous_element) == "'地址：'":
			adress += str(i.span.get_text() + '\n')
		elif repr(i.previous_element.previous_element) == "'分院：'":
			son_name += str(i.span.get_text() + '\n')
			son_url += str(i.a.get('href') + '\n')

	# print('phone-----' + phone)
	# print('adress-----' + adress)
	# print('son_name-----' + son_name)
	# print('son_url-----' + son_url)

	#简介、科研成果、获奖荣誉和先进设备放一起处理
	detail = ''
	scientific = ''
	prize = ''
	equip = ''
	detail_list = soup.select('dl[class="sur-info2 clearfix"] dd')
	for i in detail_list:
		# print(repr(i.previous_element.previous_element))

		if repr(i.previous_element.previous_element) == "'简介'":
			# print(i.get_text())
			detail += str(i.get_text() + '\n')
		elif repr(i.previous_element.previous_element) == "'科研成果'":
			# print(i.get_text())
			scientific += str(i.get_text() + '\n')
		elif repr(i.previous_element.previous_element) == "'获奖荣誉'":
			# print(i.get_text())
			prize += str(i.get_text() + '\n')
		elif repr(i.previous_element.previous_element) == "'先进设备'":
			# print(i.get_text())
			equip += str(i.get_text() + '\n')		

	# print('detail-----' + detail)
	# print('scientific-----' + scientific)
	# print('prize-----' + prize)
	# print('equip-----' + equip)
	return(level,tag,nick,phone,adress,son_name,son_url,detail,scientific,prize,equip)


# 传入url page_number 和excel对象 获取页面数据并记录
def get_page_content(url,page_number,excel_data):
	headers = {
		'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/78.0.3904.97 Safari/537.36',
		'Host': 'yyk.39.net',
		'Upgrade-Insecure-Requests': '1',
		'Cookie': 'JSESSIONID=abcZ3TdW4CbiabeOULg6w; Hm_lvt_2e44bf94e67d57ced8420d8af730dd64=1574216073; hisHos=22826%2322826; Hm_lpvt_2e44bf94e67d57ced8420d8af730dd64=1574216075',
		'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3',
		'Accept-Encoding': 'gzip, deflate',
		'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8',
		'Cache-Control': 'max-age=0',
		'Connection': 'keep-alive'
		}		 
	html_content = requests.get(url, headers=headers, timeout=120)
	html_content_text = html_content.text.strip()
	html_content_text = html_content_text.replace('\n','')
	soup = BeautifulSoup(html_content_text,'lxml')
	li_list = soup.select('li div[class="yy-msg"]')
	if len(li_list) == 0:
		with open(f'errorlog.txt', 'a+', encoding='utf-8') as f:
			f.write(f'抓取 {page_number} 数据失败\n')
	dict = {} 
	for div in li_list:
		div_parent = div.parent()		
		hospital_url = str(div_parent[0].get('href'))
		hospital_url_list1 = hospital_url.split('/')
		hospital_url_list2 = hospital_url_list1[-1].split('.')
		dict['hospital_code'] = hospital_url_list2[0]
		dict['hospital_url'] = hospital_url
		dict['hospital_icon'] = str(div_parent[1].get('src'))
		dict['hospital_name'] = str(div_parent[2].a.get_text())
		# print(dict)
		# 创建头像目录
		# os.makedirs(f'hospital_pic/',exist_ok=True)
		# 下载医生头像
		# download_hospital_icon(f'hospital_pic/', dict['hospital_icon'], dict['hospital_code'])
		# 获取医院信息
		intro_list = get_hospital_info(hospital_url)

		sheet = get_sheet_object(page_number,excel_data)
		time.sleep(1)

		title_string = ['编码', '名称', '主页', '等级', '标签', '别称', '联系方式', '地址', '分院名称', '分院主页', '简介', '科研成果', '获奖荣誉', '先进设备'];

		doctor_string = [dict['hospital_code'],dict['hospital_name'],dict['hospital_url'],intro_list[0],intro_list[1],intro_list[2],intro_list[3],intro_list[4],intro_list[5],intro_list[6],intro_list[7],intro_list[8],intro_list[9],intro_list[10]]

		current_row = sheet.max_row + 1
		for index in range(len(doctor_string)):
			sheet.cell(row=current_row,column=index+1).value = doctor_string[index]


		dict.clear()

	print('page ' + str(page_number) + ' over')
	excel_data.save("hospital_with_intro.xlsx")

		
	
# 传入page_number和excel对象 拼接url并处理
def get_page_url(page_number,excel_data):
	url = "http://yyk.39.net/hospitals/xinlike/c_p{}".format(str(page_number))
	get_page_content(url,page_number,excel_data)


# 爬取医院列表
def get_hospitals():
	try:
		excel_data = openpyxl.load_workbook("hospital_with_intro.xlsx")
	except Exception as e:
		excel_data = openpyxl.Workbook()

	min_number = 1
	max_number = 124
	index = 0

	try:
		for number in range(min_number, max_number+1):
			index += 1
			get_page_url(number,excel_data)
			if index %3 == 0:
				time.sleep(2)
	except Exception as e:
		print(f'错误代码：{e}')
		excel_data.save("hospital_with_intro.xlsx")

	excel_data.save("hospital_with_intro.xlsx")

# 获取单个医院信息
def get_one_hospital(hospital_url):
	try:
		excel_data = openpyxl.load_workbook("hospital_with_intro.xlsx")
	except Exception as e:
		excel_data = openpyxl.Workbook()

	hospital_url_list1 = hospital_url.split('/')
	hospital_url_list2 = hospital_url_list1[-1].split('.')
	# 获取医院信息
	intro_list = get_hospital_info(hospital_url)
	sheet = get_sheet_object(151,excel_data)
	title_string = ['编码', '名称', '主页', '等级', '标签', '别称', '联系方式', '地址', '分院名称', '分院主页', '简介', '科研成果', '获奖荣誉', '先进设备'];
	doctor_string = [hospital_url_list2[0],'中国人民解放军705医院',hospital_url,intro_list[0],intro_list[1],intro_list[2],intro_list[3],intro_list[4],intro_list[5],intro_list[6],intro_list[7],intro_list[8],intro_list[9],intro_list[10]]
	current_row = sheet.max_row + 1
	for index in range(len(doctor_string)):
		sheet.cell(row=current_row,column=index+1).value = doctor_string[index]

	excel_data.save("hospital_with_intro.xlsx")

if __name__ == "__main__":
	socket.setdefaulttimeout(60)
	get_hospitals()

	# get_one_hospital("http://yyk.39.net/sy/zonghe/1c8a0.html")
