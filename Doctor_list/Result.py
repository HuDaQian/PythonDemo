import openpyxl
import requests
from bs4 import BeautifulSoup
import time
import os

# 6 7 10

# 获取excel中的sheet
def get_sheet_object(sheet_index,excel_data):
	# excel_data = openpyxl.load_workbook("doctors_with_intro.xlsx")
	sheet_page_index = int(sheet_index/100) + 1;
	# print(sheet_page_index)

	try:
		return excel_data[str("Sheet"+str(sheet_page_index))]
	except Exception as e:
		excel_sheet = excel_data.create_sheet("Sheet"+str(sheet_page_index))
		title_string = ['编码', '名称', '主页', '头像', '擅长领域', '执业经历', '医院名称', '医院地址', '科室名称', '科室地址', '科室2名称', '科室2地址']
		for index in range(len(title_string)):
			excel_sheet.cell(row=1,column=index+1).value = title_string[index]
		return excel_sheet		

# 传入page_number和excel对象 拼接url并处理
def get_page_url(page_number,excel_data):
	url = "http://yyk.39.net/doctors/xinlike/c_p{}".format(str(page_number))
	get_page_content(url,page_number,excel_data)

# 传入url page_number 和excel对象 获取页面数据并记录
def get_page_content(url,page_number,excel_data):
	headers = {
	'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
		  'AppleWebKit/537.36 (KHTML, like Gecko) '
		  'Chrome/67.0.3396.99 Safari/537.36'
	}		 
	html_content = requests.get(url,headers=headers)
	html_content_text = html_content.text.strip()
	html_content_text = html_content_text.replace('\n','')
	soup = BeautifulSoup(html_content_text,'lxml')
	li_list = soup.select('li div[class="ys-msg"]')
	if len(li_list) == 0:
		with open(f'doctor_pic/errorlog.txt', 'a+', encoding='utf-8') as f:
			f.write(f'抓取 {page_number} 数据失败\n')
	dict = {} 
	for div in li_list:
		div_parent = div.parent()	
		docotrs_user_url = str(div_parent[0].get('href'))
		docotrs_user_url_list1 = docotrs_user_url.split('/')
		docotrs_user_url_list2 = docotrs_user_url_list1[-1].split('.')

		dict['user_code'] = str(docotrs_user_url_list2[0])
		dict['user_url'] = docotrs_user_url
		dict['user_url_name'] = str(div_parent[0].get('title'))
		dict['user_icon'] = str(div_parent[1].get('src'))
		dict['user_icon_name'] = str(div_parent[1].get('alt'))
		dict['user_url2'] = str(div_parent[4].get('href'))
		dict['user_url2_name'] = str(div_parent[4].get('title'))
		dict['user_hospital_url'] = '' if(str(div_parent[7].get('href')) is None) else str(div_parent[7].get('href'))
		dict['user_hospital_name'] = '' if(str(div_parent[7].get('title')) is None) else str(div_parent[7].get('title'))
		dict['user_department_url'] = '' if(str(div_parent[10].get('href')) is None) else str(div_parent[10].get('href'))
		dict['user_department_name'] = '' if(str(div_parent[10].get('title')) is None) else str(div_parent[10].get('title'))
		dict['user_department2_url'] = '' if(str(div_parent[12].get('href')) is None)  else str(div_parent[12].get('href'))
		dict['user_department2_name'] = '' if(str(div_parent[12].get('title')) is None) else str(div_parent[12].get('title'))
		# print(dict)
		# 创建头像目录
		os.makedirs(f'doctor_pic/',exist_ok=True)
		# 下载医生头像
		download_doctor_icon(f'doctor_pic', dict['user_icon'], dict['user_code'], page_number)
		# 获取医生简介
		intro_list = get_doctor_info(dict['user_url'])
		time.sleep(1)
		sheet = get_sheet_object(page_number,excel_data)
		# excel_sheet.append('编码', '名称', '主页', '头像', '擅长领域', '执业经历', '医院名称', '医院地址', '科室名称', '科室地址', '科室2名称', '科室2地址')
		doctor_string = [dict['user_code'],dict['user_icon_name'],dict['user_url'],dict['user_icon'],intro_list[0],intro_list[1],dict['user_hospital_name'],dict['user_hospital_url'],dict['user_department_name'],dict['user_department_url'],dict['user_department2_name'],dict['user_department2_url']]
		current_row = sheet.max_row + 1
		for index in range(len(doctor_string)):
			sheet.cell(row=current_row,column=index+1).value = doctor_string[index]

		dict.clear()
	print('page ' + str(page_number) + ' over')
	# time.sleep(0.5)

# 获取医生简介
def get_doctor_info(doctor_url):
	headers = {
	'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
		  'AppleWebKit/537.36 (KHTML, like Gecko) '
		  'Chrome/67.0.3396.99 Safari/537.36'
	}		 
	html_content = requests.get(doctor_url, timeout=10, headers=headers)
	html_content_text = html_content.text.strip()
	html_content_text = html_content_text.replace('\n','')
	soup = BeautifulSoup(html_content_text,'lxml')
	intro_list = soup.select('div[class="intro_more"] p')
	history_list = soup.select('div[class="hos-guide-box1"] p')
	history_string = ''
	for i in history_list:
		history_string += str(i.get_text() + '\n')
	if len(intro_list) > 0:
		intro_string = intro_list[0].get_text().strip()
		return ((intro_string[5:] if (len(intro_string) > 5) else intro_string), history_string)
	else:
		intro_more_list = soup.select('dl dd')
		if len(intro_more_list) > 2:
			intro_more_string = intro_more_list[2].get_text().strip()
		else:
			intro_more_string = ''
		intro_more_string = intro_more_string.replace('\t','')
		return ((intro_more_string[5:] if (len(intro_more_string) > 5) else intro_more_string),history_string)

# 下载医生图片
def download_doctor_icon(path, img_url, img_code, page_number):
	try:
		headers = {
			'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
			'AppleWebKit/537.36 (KHTML, like Gecko) '
			'Chrome/67.0.3396.99 Safari/537.36'
			}		 
		r = requests.get(img_url, timeout=10, headers=headers)
		print(f'下载{img_code}图片成功！')
		with open(f'{path}/{img_code}-{page_number}.png', 'wb') as f:
			f.write(r.content)
	except Exception as e:
		print(f'下载{img_code}图片失败！')
		print(f'错误代码：{e}')
		with open(f'{path}/errorlog.txt', 'a+', encoding='utf-8') as f:
			f.write(f'错误代码：{e}---下载 {img_url} 图片失败\n')

# 爬取医生列表
def get_doctors():
	try:
		excel_data = openpyxl.load_workbook("doctors_with_intro.xlsx")
	except Exception as e:
		excel_data = openpyxl.Workbook()

	min_number = 1
	max_number = 680


	try:
		for number in range(min_number, max_number+1):
			get_page_url(number,excel_data)
	except Exception as e:
		print(f'错误代码：{e}')
		excel_data.save("doctors_with_intro.xlsx")

	excel_data.save("doctors_with_intro.xlsx")


if __name__ == "__main__":
		get_doctors()
